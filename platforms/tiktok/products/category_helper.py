# -*- coding: utf-8 -*-
"""TikTok 类目：拉取、映射、Excel 下拉。"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

SCRIPT_DIR = Path(__file__).resolve().parent
CATEGORIES_JSON = SCRIPT_DIR / "categories_ph.json"
DEFAULT_EXCEL = Path(r"C:\Users\Administrator\Desktop\TikTok商品批量创建模板.xlsx")
CATEGORY_SHEET = "类目对照"
CATEGORY_COL_OLD = "类目ID"
CATEGORY_COL_NEW = "类目"

# Excel 表头颜色说明
FILL_REQUIRED = PatternFill("solid", fgColor="C00000")   # 红 = 必填
FILL_OPTIONAL = PatternFill("solid", fgColor="4472C4")   # 蓝 = 选填
FILL_CONDITIONAL = PatternFill("solid", fgColor="ED7D31")  # 橙 = 多规格时必填
FILL_AUTO = PatternFill("solid", fgColor="808080")       # 灰 = 系统自动
HEADER_FONT = Font(color="FFFFFF", bold=True)

PRODUCT_HEADERS = [
    "商品编号",
    "店铺",
    "商品标题",
    "商品描述",
    "类目",
    "品牌ID",
    "主图路径",
    "副图路径",
    "属性文件",
    "仓库ID",
    "保存方式",
    "货到付款",
    "长cm",
    "宽cm",
    "高cm",
    "重量kg",
    "是否创建",
    "创建结果",
]
PRODUCT_REQUIRED = {"商品编号", "商品标题", "类目", "主图路径"}
PRODUCT_AUTO = {"创建结果"}

SKU_HEADERS = [
    "商品编号",
    "商家SKU",
    "价格",
    "库存",
    "规格名1",
    "规格值1",
    "规格名2",
    "规格值2",
    "规格名3",
    "规格值3",
]
SKU_REQUIRED = {"商品编号", "商家SKU", "价格", "库存"}
SKU_CONDITIONAL = {"规格名1", "规格值1", "规格名2", "规格值2", "规格名3", "规格值3"}


def normalize_header(name) -> str:
    return str(name or "").strip().rstrip("*").strip()


def style_sheet_headers(
    ws,
    headers: list[str],
    *,
    required: set[str] | None = None,
    conditional: set[str] | None = None,
    auto: set[str] | None = None,
) -> None:
    required = required or set()
    conditional = conditional or set()
    auto = auto or set()
    for col, name in enumerate(headers, 1):
        base = normalize_header(name)
        if base in required:
            fill = FILL_REQUIRED
            display = f"{base}*"
        elif base in conditional:
            fill = FILL_CONDITIONAL
            display = base
        elif base in auto:
            fill = FILL_AUTO
            display = base
        else:
            fill = FILL_OPTIONAL
            display = base
        cell = ws.cell(row=1, column=col, value=display)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(12, len(display) + 2)


def apply_header_legend(ws_help) -> None:
    lines = [
        "TikTok 商品批量创建模板",
        "",
        "【表头颜色】",
        "  红色 *  = 必填",
        "  橙色    = 多规格时必填（单 SKU 可留空）",
        "  蓝色    = 选填（不填会用默认值）",
        "  灰色    = 系统自动填写，不用管",
        "",
        "【商品表 — 必填】商品编号*、商品标题*、类目*、主图路径*",
        "【商品表 — 选填】副图路径(多张用分号分隔)、店铺(默认TK6PH)、描述(英文)、品牌ID、",
        "              仓库ID(可空自动)、保存方式(默认AS_DRAFT)、货到付款、长宽高重量",
        "",
        "【SKU表 — 必填】商品编号*、商家SKU*、价格*、库存*",
        "【SKU表 — 多规格时填】规格名/规格值 1~3（单 SKU 整行规格列留空）",
        "",
        "1. 商品编号：同一商品的多行 SKU 用相同编号关联。",
        "2. 类目：点下拉选中文路径；下拉空则先 run_sync_categories.bat。",
        "3. 填好后：python run_excel_create.py --execute",
    ]
    for i, line in enumerate(lines, 1):
        ws_help.cell(row=i, column=1, value=line)
    ws_help.column_dimensions["A"].width = 100


def build_category_catalog(categories: list[dict]) -> list[dict]:
    by_id = {str(c.get("id") or ""): c for c in categories}

    def path_for(category_id: str) -> str:
        parts: list[str] = []
        cur = by_id.get(category_id)
        seen: set[str] = set()
        while cur and str(cur.get("id") or "") not in seen:
            cid = str(cur.get("id") or "")
            seen.add(cid)
            parts.append(str(cur.get("local_name") or cid))
            parent_id = str(cur.get("parent_id") or "0")
            if parent_id in ("0", ""):
                break
            cur = by_id.get(parent_id)
        return " > ".join(reversed(parts))

    rows: list[dict] = []
    for c in categories:
        if not c.get("is_leaf"):
            continue
        cid = str(c.get("id") or "")
        path = path_for(cid)
        name = str(c.get("local_name") or "")
        label = f"{cid} | {path}"
        rows.append(
            {
                "id": cid,
                "name": name,
                "path": path,
                "label": label,
            }
        )
    rows.sort(key=lambda x: (x["path"].lower(), x["id"]))
    return rows


def save_category_catalog(rows: list[dict], shop: str = "TK6PH", region: str = "PH", locale: str = "zh-CN") -> None:
    payload = {
        "shop": shop,
        "region": region,
        "locale": locale,
        "count": len(rows),
        "categories": rows,
        "by_id": {r["id"]: r for r in rows},
        "by_label": {r["label"]: r for r in rows},
    }
    CATEGORIES_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def load_category_catalog() -> dict:
    if not CATEGORIES_JSON.exists():
        return {"categories": [], "by_id": {}, "by_label": {}}
    data = json.loads(CATEGORIES_JSON.read_text(encoding="utf-8"))
    if "by_id" not in data:
        rows = data.get("categories") or []
        data["by_id"] = {str(r["id"]): r for r in rows}
        data["by_label"] = {str(r["label"]): r for r in rows}
    return data


def resolve_category_id(
    raw: str,
    *,
    catalog: dict | None = None,
) -> str:
    text = str(raw or "").strip()
    if not text:
        return ""
    if text.isdigit():
        return text
    if " | " in text:
        left = text.split(" | ", 1)[0].strip()
        if left.isdigit():
            return left
    cat = catalog or load_category_catalog()
    by_label = cat.get("by_label") or {}
    if text in by_label:
        return str(by_label[text]["id"])
    by_id = cat.get("by_id") or {}
    if text in by_id:
        return text
    # 按路径或名称模糊匹配（取第一个）
    lower = text.lower()
    for row in cat.get("categories") or []:
        if lower == str(row.get("name") or "").lower():
            return str(row["id"])
        if lower in str(row.get("path") or "").lower():
            return str(row["id"])
    return text if text.isdigit() else ""


def apply_category_sheet(wb, rows: list[dict]) -> None:
    if CATEGORY_SHEET in wb.sheetnames:
        del wb[CATEGORY_SHEET]
    ws = wb.create_sheet(CATEGORY_SHEET)
    headers = ["类目ID", "类目名称", "完整路径", "下拉选项"]
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    for row_idx, item in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=item["id"])
        ws.cell(row=row_idx, column=2, value=item["name"])
        ws.cell(row=row_idx, column=3, value=item["path"])
        ws.cell(row=row_idx, column=4, value=item["label"])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 70
    ws.freeze_panes = "A2"

    if "商品" not in wb.sheetnames:
        return
    ws_prod = wb["商品"]
    headers = [ws_prod.cell(row=1, column=c).value for c in range(1, ws_prod.max_column + 1)]
    cat_col = None
    for i, name in enumerate(headers, 1):
        if normalize_header(name) in (CATEGORY_COL_NEW, CATEGORY_COL_OLD):
            cat_col = i
            ws_prod.cell(row=1, column=i, value=f"{CATEGORY_COL_NEW}*")
            ws_prod.cell(row=1, column=i).fill = FILL_REQUIRED
            ws_prod.cell(row=1, column=i).font = HEADER_FONT
            break
    if cat_col is None:
        return

    last_row = max(2, len(rows) + 1)
    formula = f"'{CATEGORY_SHEET}'!$D$2:$D${last_row}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "请从下拉列表选择类目"
    dv.errorTitle = "类目无效"
    dv.prompt = "选择与卖家后台一致的商品类目"
    dv.promptTitle = "选择类目"
    ws_prod.add_data_validation(dv)
    for r in range(2, 1001):
        dv.add(ws_prod.cell(row=r, column=cat_col))


def default_category_label(category_id: str = "810376") -> str:
    cat = load_category_catalog()
    row = (cat.get("by_id") or {}).get(str(category_id))
    if row:
        return row["label"]
    return str(category_id)
