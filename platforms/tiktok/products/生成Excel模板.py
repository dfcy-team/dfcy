# -*- coding: utf-8 -*-
"""生成桌面 Excel 批量创建模板。"""

from pathlib import Path

from openpyxl import Workbook

from category_helper import (
    DEFAULT_EXCEL,
    PRODUCT_AUTO,
    PRODUCT_HEADERS,
    PRODUCT_REQUIRED,
    SKU_CONDITIONAL,
    SKU_HEADERS,
    SKU_REQUIRED,
    apply_category_sheet,
    apply_header_legend,
    default_category_label,
    load_category_catalog,
    style_sheet_headers,
)

OUT = DEFAULT_EXCEL

EXAMPLE_PRODUCT = [
    1,
    "TK6PH",
    "Velvet Sofa Cover API Test Multi SKU",
    "<p>API test product. In stock, ships within 24h.</p>",
    default_category_label("810376"),
    "7611082281133836033",
    r"samples\main.jpg",
    r"samples\sub1.jpg;samples\sub2.jpg",
    "auto by category",
    "7432356829526001413",
    "AS_DRAFT",
    1,
    10,
    10,
    5,
    0.3,
    1,
    "",
]

EXAMPLE_SKUS = [
    [1, "API-EXCEL-BK-70", 99, 5, "Color", "Black", "Size", "70x70cm", "", ""],
    [1, "API-EXCEL-BK-90", 109, 8, "Color", "Black", "Size", "90x90cm", "", ""],
    [1, "API-EXCEL-GY-70", 99, 6, "Color", "Grey", "Size", "70x70cm", "", ""],
]


def main():
    wb = Workbook()
    ws_help = wb.active
    ws_help.title = "填写说明"
    apply_header_legend(ws_help)

    ws_prod = wb.create_sheet("商品")
    style_sheet_headers(
        ws_prod,
        PRODUCT_HEADERS,
        required=PRODUCT_REQUIRED,
        auto=PRODUCT_AUTO,
    )
    for col, val in enumerate(EXAMPLE_PRODUCT, 1):
        ws_prod.cell(row=2, column=col, value=val)

    ws_sku = wb.create_sheet("SKU")
    style_sheet_headers(
        ws_sku,
        SKU_HEADERS,
        required=SKU_REQUIRED,
        conditional=SKU_CONDITIONAL,
    )
    for row_idx, row in enumerate(EXAMPLE_SKUS, 2):
        for col_idx, val in enumerate(row, 1):
            ws_sku.cell(row=row_idx, column=col_idx, value=val)

    catalog = load_category_catalog()
    rows = catalog.get("categories") or []
    if rows:
        apply_category_sheet(wb, rows)

    wb.save(OUT)
    print(f"已生成: {OUT}")
    if not rows:
        print("提示: 类目下拉尚未填充，请运行 run_sync_categories.bat")


if __name__ == "__main__":
    main()
