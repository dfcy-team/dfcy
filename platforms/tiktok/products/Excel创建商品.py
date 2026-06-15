# -*- coding: utf-8 -*-
"""从桌面 Excel 批量创建 TikTok 商品。"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from category_helper import load_category_catalog, resolve_category_id
from product_api import (
    attributes_template_from_api,
    build_create_product_body,
    create_product,
    default_currency,
    default_sales_warehouse_id,
    get_category_attributes,
    get_product,
    ini_get,
    load_ini,
    load_json_file,
    parse_product_attributes,
    parse_image_path_list,
    print_api_result,
    product_status_label,
    recommend_category,
    save_json,
    setup_client,
    sku_inventory_hint,
    upload_product_image,
    upload_product_images,
)

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL = Path(r"C:\Users\Administrator\Desktop\TikTok商品批量创建模板.xlsx")


def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def cell_bool(value, default: bool = True) -> bool:
    raw = cell_str(value).lower()
    if not raw:
        return default
    return raw in ("1", "true", "yes", "y", "是", "on")


def read_products_sheet(ws) -> dict[str, dict]:
    headers = [cell_str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {cell_str(h).rstrip("*").strip(): i for i, h in enumerate(headers)}
    cat_col = "类目" if "类目" in idx else "类目ID"
    required = ["商品编号", "商品标题", cat_col, "主图路径"]
    for col in required:
        if col not in idx:
            raise ValueError(f"「商品」表缺少列: {col}")

    products: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        pid = cell_str(row[idx["商品编号"]])
        if not pid:
            continue
        products[pid] = {
            "product_no": pid,
            "shop": cell_str(row[idx.get("店铺", -1)]) if "店铺" in idx else "",
            "title": cell_str(row[idx["商品标题"]]),
            "description": cell_str(row[idx.get("商品描述", -1)]) if "商品描述" in idx else "",
            "category_raw": cell_str(row[idx[cat_col]]),
            "category_id": "",
            "brand_id": cell_str(row[idx.get("品牌ID", -1)]) if "品牌ID" in idx else "",
            "image_path": cell_str(row[idx["主图路径"]]),
            "sub_image_paths": cell_str(row[idx.get("副图路径", -1)]) if "副图路径" in idx else "",
            "attributes_json": cell_str(row[idx.get("属性文件", -1)]) if "属性文件" in idx else "",
            "warehouse_id": cell_str(row[idx.get("仓库ID", -1)]) if "仓库ID" in idx else "",
            "save_mode": cell_str(row[idx.get("保存方式", -1)]) if "保存方式" in idx else "AS_DRAFT",
            "is_cod_allowed": cell_bool(row[idx.get("货到付款", -1)] if "货到付款" in idx else "", True),
            "package_length": cell_str(row[idx.get("长cm", -1)]) if "长cm" in idx else "10",
            "package_width": cell_str(row[idx.get("宽cm", -1)]) if "宽cm" in idx else "10",
            "package_height": cell_str(row[idx.get("高cm", -1)]) if "高cm" in idx else "5",
            "package_weight": cell_str(row[idx.get("重量kg", -1)]) if "重量kg" in idx else "0.3",
            "enabled": cell_bool(row[idx.get("是否创建", -1)] if "是否创建" in idx else "1", True),
            "result_col": idx.get("创建结果"),
            "row_no": ws.max_row,
        }
    # fix row numbers
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if not row or not any(c.value for c in row):
            continue
        pid = cell_str(row[idx["商品编号"]].value)
        if pid in products:
            products[pid]["row_no"] = row_idx
            products[pid]["result_cell"] = row[idx["创建结果"]].column if "创建结果" in idx else None
    return products


def read_skus_sheet(ws) -> dict[str, list[dict]]:
    headers = [cell_str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {cell_str(h).rstrip("*").strip(): i for i, h in enumerate(headers)}
    if "商品编号" not in idx or "商家SKU" not in idx:
        raise ValueError("「SKU」表缺少列: 商品编号 / 商家SKU")

    grouped: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        product_no = cell_str(row[idx["商品编号"]])
        seller_sku = cell_str(row[idx["商家SKU"]])
        if not product_no or not seller_sku:
            continue
        sku = {
            "seller_sku": seller_sku,
            "price": cell_str(row[idx.get("价格", -1)]) or "99",
            "quantity": int(float(cell_str(row[idx.get("库存", -1)]) or "1")),
            "sales_attributes": [],
        }
        for n in (1, 2, 3):
            name_key = f"规格名{n}"
            value_key = f"规格值{n}"
            if name_key not in idx or value_key not in idx:
                continue
            name = cell_str(row[idx[name_key]])
            value = cell_str(row[idx[value_key]])
            if not name and not value:
                continue
            attr: dict = {"name": name or value_key, "value_name": value}
            if name.lower() == "color" or name == "Color" or name == "颜色":
                attr["id"] = "100000"
                attr["name"] = "Color"
            sku["sales_attributes"].append(attr)
        grouped.setdefault(product_no, []).append(sku)
    return grouped


def resolve_image_paths(raw: str, base_resolver) -> list[Path]:
    out: list[Path] = []
    for item in parse_image_path_list(raw):
        if item.lower().startswith("tos-"):
            continue
        out.append(base_resolver(item))
    return out


def resolve_image_path(raw: str) -> Path:
    p = Path(raw)
    if p.is_absolute() and p.is_file():
        return p
    for base in (SCRIPT_DIR, Path(r"C:\Users\Administrator\Desktop"), SCRIPT_DIR.parent.parent):
        candidate = base / raw
        if candidate.is_file():
            return candidate
    return SCRIPT_DIR / raw


def upload_all_product_images(
    client,
    token: str,
    cipher: str,
    main_path: str,
    sub_paths: str,
    *,
    dry_run: bool,
) -> list[str]:
    main_raw = (main_path or "").strip()
    if main_raw.lower().startswith("tos-"):
        uris = [main_raw]
    else:
        main_p = resolve_image_path(main_raw)
        if not main_p.is_file():
            raise FileNotFoundError(f"主图不存在: {main_p}")
        if dry_run:
            print(f"[dry_run] 主图: {main_p}")
            uris = ["DRY_RUN_MAIN_URI"]
        else:
            print(f"上传主图: {main_p}")
            uris = [upload_product_image(client, token, cipher, main_p)]

    sub_list = resolve_image_paths(sub_paths, resolve_image_path)
    if sub_list:
        missing = [p for p in sub_list if not p.is_file()]
        if missing:
            raise FileNotFoundError(f"副图不存在: {missing[0]}")
        if dry_run:
            print(f"[dry_run] 副图 x{len(sub_list)}: " + "; ".join(str(p) for p in sub_list))
            uris.extend([f"DRY_RUN_SUB_URI_{i}" for i in range(1, len(sub_list) + 1)])
        else:
            print(f"上传副图 x{len(sub_list)}")
            uris.extend(upload_product_images(client, token, cipher, sub_list))
    return uris


def resolve_attributes_file(
    name: str,
    *,
    category_id: str,
    client,
    token: str,
    cipher: str,
) -> list[dict]:
    skip_auto = {"留空=自动按类目生成", "自动", "auto", "auto by category"}
    fname = (name or "").strip()
    if fname and fname.lower() not in skip_auto:
        path = SCRIPT_DIR / fname
        if path.exists():
            return parse_product_attributes(load_json_file(path))
        raise FileNotFoundError(f"属性文件不存在: {path}")

    per_cat = SCRIPT_DIR / f"category_{category_id}_attrs.json"
    if per_cat.exists():
        raw = load_json_file(per_cat)
        if isinstance(raw, dict):
            return parse_product_attributes(raw.get("product_attributes") or raw)
        return parse_product_attributes(raw)

    attrs = get_category_attributes(client, token, cipher, category_id)
    template = attributes_template_from_api(attrs)
    per_cat.write_text(
        json.dumps({"category_id": category_id, "product_attributes": template}, ensure_ascii=False, indent=2)
        + "\n",
        encoding="utf-8",
    )
    print(f"已自动生成属性模板: {per_cat.name}（每属性取第一个可选值，可按需编辑）")
    return template


def build_skus_for_api(sku_rows: list[dict], warehouse_id: str) -> list[dict]:
    cur = default_currency()
    out = []
    for item in sku_rows:
        sku = {
            "seller_sku": item["seller_sku"],
            "price": {"amount": str(item["price"]), "currency": cur},
            "inventory": [{"warehouse_id": warehouse_id, "quantity": int(item["quantity"])}],
        }
        if item["sales_attributes"]:
            sku["sales_attributes"] = item["sales_attributes"]
        out.append(sku)
    return out


def main() -> int:
    cp = load_ini()
    ap = argparse.ArgumentParser(description="从 Excel 批量创建 TikTok 商品")
    default_excel = ini_get(cp, "excel_create", "excel_path", str(DEFAULT_EXCEL)) or str(DEFAULT_EXCEL)
    ap.add_argument("--excel", default=default_excel, help="Excel 路径")
    ap.add_argument("--shop", "-s", default=ini_get(cp, "common", "shop", "TK6PH"))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--execute", action="store_true")
    args = ap.parse_args()
    dry_run = args.dry_run and not args.execute

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"找不到 Excel: {excel_path}")
        print("请先在桌面填写: TikTok商品批量创建模板.xlsx")
        return 1

    wb = load_workbook(excel_path)
    if "商品" not in wb.sheetnames or "SKU" not in wb.sheetnames:
        print("Excel 需包含「商品」「SKU」两个工作表")
        return 1

    products = read_products_sheet(wb["商品"])
    skus = read_skus_sheet(wb["SKU"])
    if not products:
        print("「商品」表没有有效数据")
        return 1

    argv = ["--shop", args.shop] if args.shop else []
    client, token, cipher, cfg_path = setup_client(argv)
    print(f"店铺: {cfg_path.name}")
    print(f"Excel: {excel_path}")

    default_brand = ini_get(cp, "create", "brand_id", "")
    default_wh = ini_get(cp, "create", "warehouse_id", "")
    if not default_wh:
        default_wh = default_sales_warehouse_id(client, token, cipher) or ""
    category_catalog = load_category_catalog()

    exit_code = 0
    results = []

    for product_no, prod in products.items():
        if not prod["enabled"]:
            print(f"\n[跳过] 商品编号 {product_no}（是否创建=0）")
            continue

        category_id = resolve_category_id(prod["category_raw"], catalog=category_catalog)
        if not category_id and prod["title"]:
            rr = recommend_category(client, token, cipher, prod["title"], prod["description"])
            if rr.get("code") == 0:
                category_id = str((rr.get("data") or {}).get("leaf_category_id") or "")
                if category_id:
                    print(f"[推荐类目] {prod['title'][:40]} -> {category_id}")
        prod["category_id"] = category_id

        sku_rows = skus.get(product_no) or []
        if not sku_rows:
            msg = "缺少 SKU 行"
            print(f"\n[失败] {product_no} {prod['title']} — {msg}")
            if prod.get("result_cell"):
                wb["商品"].cell(row=prod["row_no"], column=prod["result_cell"], value=f"失败: {msg}")
            exit_code = 1
            continue

        title = prod["title"]
        if not title or not prod["category_id"]:
            msg = "标题为空，或类目无法识别（请从下拉选择类目）"
            print(f"\n[失败] {product_no} — {msg}")
            exit_code = 1
            continue

        print(f"类目 ID = {prod['category_id']}")

        warehouse_id = prod["warehouse_id"] or default_wh
        if not warehouse_id:
            msg = "无仓库ID"
            print(f"\n[失败] {product_no} {title} — {msg}")
            exit_code = 1
            continue

        print(f"\n--- 商品编号 {product_no}: {title} ({len(sku_rows)} SKU) ---")

        try:
            image_uris = upload_all_product_images(
                client,
                token,
                cipher,
                prod["image_path"],
                prod.get("sub_image_paths") or "",
                dry_run=dry_run,
            )
        except FileNotFoundError as e:
            msg = str(e)
            print(f"[失败] {msg}")
            if prod.get("result_cell"):
                wb["商品"].cell(row=prod["row_no"], column=prod["result_cell"], value=f"失败: {msg}")
            exit_code = 1
            continue

        try:
            product_attributes = resolve_attributes_file(
                prod["attributes_json"],
                category_id=prod["category_id"],
                client=client,
                token=token,
                cipher=cipher,
            )
        except Exception as e:
            msg = str(e)
            print(f"[失败] {msg}")
            if prod.get("result_cell"):
                wb["商品"].cell(row=prod["row_no"], column=prod["result_cell"], value=f"失败: {msg}")
            exit_code = 1
            continue

        description = prod["description"] or f"<p>{title}</p>"
        brand_id = prod["brand_id"] or default_brand
        api_skus = build_skus_for_api(sku_rows, warehouse_id)

        body = build_create_product_body(
            title=title,
            description=description,
            category_id=prod["category_id"],
            main_image_uri=image_uris[0] if image_uris else "",
            main_image_uris=image_uris,
            product_attributes=product_attributes,
            seller_sku=sku_rows[0]["seller_sku"],
            price_amount=str(sku_rows[0]["price"]),
            warehouse_id=warehouse_id,
            quantity=int(sku_rows[0]["quantity"]),
            brand_id=brand_id,
            save_mode=prod["save_mode"] or "AS_DRAFT",
            is_cod_allowed=prod["is_cod_allowed"],
            package_length=prod["package_length"],
            package_width=prod["package_width"],
            package_height=prod["package_height"],
            package_weight=prod["package_weight"],
            skus=api_skus,
        )

        save_json(f"excel_create_{product_no}_request.json", body)

        if dry_run:
            print("[dry_run] 请求体已保存，未调用 API")
            print(json.dumps(body, ensure_ascii=False, indent=2)[:2000])
            if prod.get("result_cell"):
                wb["商品"].cell(row=prod["row_no"], column=prod["result_cell"], value="dry_run OK")
            continue

        r = create_product(client, token, cipher, body)
        print_api_result(r, "创建商品")
        save_json(f"excel_create_{product_no}_last.json", r)

        if r.get("code") != 0:
            msg = r.get("message") or "API 失败"
            if prod.get("result_cell"):
                wb["商品"].cell(row=prod["row_no"], column=prod["result_cell"], value=f"失败: {msg}")
            exit_code = 1
            continue

        data = r.get("data") or {}
        new_pid = str(data.get("product_id") or data.get("id") or "")
        status = ""
        sku_lines = []
        if new_pid:
            try:
                detail = get_product(client, token, cipher, new_pid)
                status = product_status_label(detail)
                for sku in detail.get("skus") or []:
                    wh, qty = sku_inventory_hint(sku)
                    sku_lines.append(f"{sku.get('seller_sku')}={sku.get('id')} qty{qty}")
            except Exception:
                pass

        result_text = f"OK product_id={new_pid} {status}"
        if sku_lines:
            result_text += " | " + "; ".join(sku_lines[:5])
        print(result_text)
        if prod.get("result_cell"):
            wb["商品"].cell(row=prod["row_no"], column=prod["result_cell"], value=result_text)
        results.append({"product_no": product_no, "product_id": new_pid, "status": status})

    if not dry_run:
        try:
            wb.save(excel_path)
            print(f"\n结果已写回 Excel: {excel_path}")
        except PermissionError:
            alt = excel_path.with_name(excel_path.stem + "_结果" + excel_path.suffix)
            wb.save(alt)
            print(f"\nExcel 可能正打开，结果已另存: {alt}")
    save_json("excel_create_summary.json", results)
    return exit_code


if __name__ == "__main__":
    sys.exit(main())
