# -*- coding: utf-8 -*-
"""给已有 Excel 刷必填表头颜色（不改动数据行）。"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

from category_helper import (
    DEFAULT_EXCEL,
    PRODUCT_AUTO,
    PRODUCT_HEADERS,
    PRODUCT_REQUIRED,
    SKU_CONDITIONAL,
    SKU_HEADERS,
    SKU_REQUIRED,
    apply_category_sheet,
    apply_header_legend,
    load_category_catalog,
    style_sheet_headers,
)


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_EXCEL
    if not path.exists():
        print(f"找不到: {path}")
        return 1

    wb = load_workbook(path)
    if "填写说明" in wb.sheetnames:
        apply_header_legend(wb["填写说明"])
    if "商品" in wb.sheetnames:
        style_sheet_headers(
            wb["商品"],
            PRODUCT_HEADERS,
            required=PRODUCT_REQUIRED,
            auto=PRODUCT_AUTO,
        )
    if "SKU" in wb.sheetnames:
        style_sheet_headers(
            wb["SKU"],
            SKU_HEADERS,
            required=SKU_REQUIRED,
            conditional=SKU_CONDITIONAL,
        )

    rows = (load_category_catalog().get("categories") or [])
    if rows:
        apply_category_sheet(wb, rows)

    wb.save(path)
    print(f"已标记必填项: {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
