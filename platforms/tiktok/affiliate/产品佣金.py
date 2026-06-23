# -*- coding: utf-8 -*-
"""
TikTok Shop 联盟产品佣金导出 — 对齐联盟中心「表现 → 商品」Excel

数据来源（混合）:
  GMV / 订单 / 件数 / 达人内容 → GET /analytics/202605/shop_products/performance
    权限: data.shop_analytics.public.read
  预计佣金 → POST /affiliate_seller/202410/orders/search
    权限: seller.affiliate_collaboration.read
  商品名称/类目 → seller.product.basic

用法:
  python 产品佣金.py --shop TK1PH --start 2026-06-16 --end 2026-06-16
  python 产品佣金.py --shop TK1PH --start 2026-06-16 --compare Z:\\Tk每日数据\\产品佣金\\TIKTOK1号店PH_产品佣金_2026-06-16_2026-06-16.xlsx
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
ENV_ROOT = SCRIPT_DIR.parent
_TEST_ENV = ENV_ROOT / "test_env"
_PROJECT = SCRIPT_DIR.parents[2]
if str(_TEST_ENV) not in sys.path:
    sys.path.insert(0, str(_TEST_ENV))
if str(_PROJECT) not in sys.path:
    sys.path.insert(0, str(_PROJECT))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from common.timezone_util import REGION_IANA, get_timezone
from tts_client import TikTokShopClient, cfg, get_shop_token, init_shop_config, is_ok

PRODUCTS_DIR = ENV_ROOT / "products"
if str(PRODUCTS_DIR) not in sys.path:
    sys.path.insert(0, str(PRODUCTS_DIR))
from product_api import get_product  # noqa: E402

EXCEL_HEADERS = [
    "Product name",
    "Product ID",
    "Product category",
    "Creator-attributed GMV",
    "Refunds",
    "Creator-attributed items sold",
    "Items refunded",
    "Attributed orders",
    "Avg. daily customers",
    "Avg. daily creators with sales",
    "Avg. daily creators posted content",
    "Avg. daily videos with sales",
    "Avg. daily LIVE streams with sales",
    "Videos",
    "LIVE streams",
    "Est. commission",
    "Samples shipped",
]

COMPARE_METRICS = EXCEL_HEADERS[3:]

# 联盟中心中文表头 → 英文标准列名
ZH_HEADER_MAP = {
    "商品名称": "Product name",
    "商品 ID": "Product ID",
    "商品ID": "Product ID",
    "商品类目": "Product category",
    "达人归因 GMV": "Creator-attributed GMV",
    "退款金额": "Refunds",
    "达人归因成交件数": "Creator-attributed items sold",
    "已退款的商品件数": "Items refunded",
    "已退款商品件数": "Items refunded",
    "成交订单数": "Attributed orders",
    "归因订单数": "Attributed orders",
    "日均客户数": "Avg. daily customers",
    "日均活跃达人数": "Avg. daily creators with sales",
    "已发布内容的达人数": "Avg. daily creators posted content",
    "已发布内容的日均达人数": "Avg. daily creators posted content",
    "日均有销售的视频数": "Avg. daily videos with sales",
    "日均视频数": "Avg. daily videos with sales",
    "日均直播数": "Avg. daily LIVE streams with sales",
    "日均直播场次": "Avg. daily LIVE streams with sales",
    "视频数": "Videos",
    "直播数": "LIVE streams",
    "预计佣金": "Est. commission",
    "已发样品数": "Samples shipped",
    "已发货样品数": "Samples shipped",
}

AFFILIATE_ORDERS_PATH = "/affiliate_seller/202410/orders/search"


def infer_region() -> str:
    r = (cfg("TTS_TARGET_REGION", "") or "PH").strip().upper()
    if r in REGION_IANA:
        return r
    tag = (cfg("TTS_EXPORT_SHOP_TAG", "") or cfg("TTS_CONFIG_LABEL", "")).upper()
    for code in REGION_IANA:
        if tag.endswith(code):
            return code
    return "PH"


def parse_day(s: str) -> datetime:
    return datetime.strptime(s.strip()[:10], "%Y-%m-%d")


def api_end_date_exclusive(end: str) -> str:
    """Analytics API end_date_lt 为排他上界。"""
    d = parse_day(end) + timedelta(days=1)
    return d.strftime("%Y-%m-%d")


def money_amount(obj) -> float:
    if obj is None:
        return 0.0
    if isinstance(obj, dict):
        try:
            return float(obj.get("amount") or 0)
        except (TypeError, ValueError):
            return 0.0
    try:
        return float(obj)
    except (TypeError, ValueError):
        return 0.0


def cell_str(v) -> str:
    if v is None:
        return "0"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def day_range_unix(start: str, end: str, region: str) -> tuple[int, int]:
    tz_name = REGION_IANA.get(region, "Asia/Manila")
    tz = get_timezone(tz_name)
    t0 = int(parse_day(start).replace(tzinfo=tz).timestamp())
    t1 = int((parse_day(end) + timedelta(days=1)).replace(tzinfo=tz).timestamp())
    return t0, t1


def fetch_all_affiliate_orders(
    client: TikTokShopClient, token: str, cipher: str, start: str, end: str, region: str
) -> list[dict]:
    t0, t1 = day_range_unix(start, end, region)
    orders: list[dict] = []
    page_token: str | None = None
    for _ in range(100):
        extra: dict = {"shop_cipher": cipher, "page_size": 50}
        if page_token:
            extra["page_token"] = page_token
        r = client.post(
            AFFILIATE_ORDERS_PATH,
            token,
            {"create_time_ge": t0, "create_time_lt": t1},
            extra,
        )
        if not is_ok(r):
            code = r.get("code")
            msg = r.get("message") or ""
            if code == 105005:
                print("  [affiliate orders] 105005 — 店铺需重新授权 seller.affiliate_collaboration.read")
            else:
                print(f"  [affiliate orders] code={code} {msg[:80]}")
            break
        data = r.get("data") or {}
        batch = data.get("orders") or []
        orders.extend(batch)
        page_token = data.get("next_page_token") or data.get("page_token")
        if not batch or not page_token:
            break
    return orders


def aggregate_affiliate_orders(orders: list[dict]) -> dict[str, dict]:
    """按 product_id 汇总联盟订单 SKU（仅佣金）。"""
    out: dict[str, dict] = {}
    for order in orders:
        for sku in order.get("skus") or []:
            pid = str(sku.get("product_id") or "")
            if not pid:
                continue
            bucket = out.setdefault(pid, {"est_commission": 0.0})
            bucket["est_commission"] += money_amount(sku.get("estimated_paid_commission"))
            bucket["est_commission"] += money_amount(sku.get("estimated_paid_partner_commission"))
    return out


def fetch_all_product_performance(
    client: TikTokShopClient, token: str, cipher: str, start: str, end: str
) -> list[dict]:
    path = "/analytics/202605/shop_products/performance"
    end_lt = api_end_date_exclusive(end)
    page = 1
    out: list[dict] = []
    while True:
        extra = {
            "shop_cipher": cipher,
            "start_date_ge": start[:10],
            "end_date_lt": end_lt,
            "page_size": 100,
            "page_number": page,
        }
        r = client.get(path, token, extra)
        if not is_ok(r):
            raise RuntimeError(f"商品表现 API 失败: code={r.get('code')} {r.get('message')}")
        batch = (r.get("data") or {}).get("products") or []
        if not batch:
            break
        out.extend(batch)
        if len(batch) < 100:
            break
        page += 1
    return out


def load_product_meta(
    client: TikTokShopClient, token: str, cipher: str, product_id: str, cache: dict[str, dict]
) -> dict:
    if product_id in cache:
        return cache[product_id]
    meta = {"title": "", "category": ""}
    try:
        p = get_product(client, token, cipher, product_id)
        if p:
            meta["title"] = str(p.get("title") or p.get("product_name") or "")
            cat = p.get("category_chains") or p.get("categories") or []
            if isinstance(cat, list) and cat:
                leaf = cat[-1]
                if isinstance(leaf, dict):
                    meta["category"] = str(leaf.get("local_name") or leaf.get("name") or "")
            if not meta["category"]:
                meta["category"] = str(p.get("category_name") or "")
    except Exception:
        pass
    cache[product_id] = meta
    return meta


def build_row(
    item: dict,
    meta: dict,
    *,
    single_day: bool,
    order_agg: dict | None = None,
) -> list:
    aff = item.get("affiliate_total_performance") or {}
    vid = item.get("affiliate_video_performance") or {}
    gmv = money_amount(aff.get("attributed_gmv"))
    items_sold = aff.get("attributed_sold_items") or 0
    orders = aff.get("attributed_orders") or 0
    customers = aff.get("estimated_customers") or 0
    posted = aff.get("avg_daily_creator_posted_content") or 0
    videos = vid.get("new_video_count") or 0
    creators_sales = customers if single_day else customers

    est_commission = ""
    if order_agg:
        comm = round(float(order_agg.get("est_commission") or 0), 2)
        if comm > 0:
            est_commission = comm

    return [
        meta.get("title", ""),
        str(item.get("id") or ""),
        meta.get("category", ""),
        gmv,
        0,  # Refunds — 分析 API 无独立字段
        cell_str(items_sold),
        "0",  # Items refunded
        orders,
        customers,
        creators_sales,
        posted,
        "",  # Avg. daily videos with sales
        "",  # Avg. daily LIVE streams with sales
        videos,
        0,  # LIVE streams
        est_commission,
        "",  # Samples shipped
    ]


def affiliate_activity_score(row: list) -> float:
    try:
        gmv = float(row[3] or 0)
    except (TypeError, ValueError):
        gmv = 0.0
    posted = row[10] or 0
    videos = row[13] or 0
    try:
        posted_n = float(posted)
    except (TypeError, ValueError):
        posted_n = 0.0
    try:
        videos_n = float(videos)
    except (TypeError, ValueError):
        videos_n = 0.0
    try:
        comm = float(row[15] or 0)
    except (TypeError, ValueError):
        comm = 0.0
    return gmv + posted_n + videos_n + comm


def filter_rows_like_affiliate_center(rows: list[list]) -> list[list]:
    kept = [r for r in rows if affiliate_activity_score(r) > 0]
    kept.sort(key=lambda r: float(r[3] or 0), reverse=True)
    return kept


def export_excel(rows: list[list], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws.append(EXCEL_HEADERS)
    for row in rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def read_reference(path: Path) -> dict[str, list]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    norm_header = [ZH_HEADER_MAP.get(h, h) for h in header]
    idx_id = norm_header.index("Product ID") if "Product ID" in norm_header else 1
    out: dict[str, dict] = {}
    for row in rows[1:]:
        if not row or row[idx_id] is None:
            continue
        pid = str(row[idx_id])
        mapped: dict = {}
        for i, h in enumerate(norm_header):
            if h:
                mapped[h] = row[i] if i < len(row) else None
        out[pid] = mapped
    return out


def norm_compare(v) -> str:
    if v is None or v == "":
        return "0"
    if isinstance(v, float):
        return f"{v:.2f}".rstrip("0").rstrip(".") if v % 1 else str(int(v))
    s = str(v).strip()
    for ch in "₱$€£¥฿,":
        s = s.replace(ch, "")
    s = s.strip()
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return f"{f:.2f}".rstrip("0").rstrip(".")
    except ValueError:
        return s


def values_match(ref_v, api_v, header: str) -> bool:
    if header in ("Samples shipped",):
        if norm_compare(api_v) in ("0", "") and norm_compare(ref_v) not in ("0", ""):
            return False
    if header in (
        "Avg. daily videos with sales",
        "Avg. daily LIVE streams with sales",
        "Avg. daily creators with sales",
        "Avg. daily customers",
        "Avg. daily creators posted content",
    ):
        if norm_compare(api_v) == "" and norm_compare(ref_v) not in ("0", ""):
            return False
    return norm_compare(ref_v) == norm_compare(api_v)


def write_comparison(
    ref_path: Path,
    api_rows: list[list],
    out_path: Path,
    api_by_id: dict[str, list],
) -> dict:
    ref_map = read_reference(ref_path)
    wb = Workbook()

    ws = wb.active
    ws.title = "对比明细"
    ws.append(
        [
            "Product ID",
            "Product name(参考)",
            "字段",
            "联盟中心(参考)",
            "API导出",
            "是否一致",
            "说明",
        ]
    )
    red = PatternFill("solid", fgColor="FFC7CE")
    match_count = 0
    total_checks = 0
    missing_api = 0

    api_ids = set(api_by_id)
    ref_ids = set(ref_map)

    for pid in sorted(
        ref_ids,
        key=lambda x: -float(norm_compare(ref_map[x].get("Creator-attributed GMV") or 0)),
    ):
        ref = ref_map[pid]
        api_row = api_by_id.get(pid)
        name = ref.get("Product name") or ""
        if not api_row:
            ws.append([pid, name, "(整行)", "有", "API无此商品", "否", "参考有、API列表未返回"])
            missing_api += 1
            continue
        api_dict = dict(zip(EXCEL_HEADERS, api_row))
        for metric in COMPARE_METRICS:
            ref_v = ref.get(metric)
            api_v = api_dict.get(metric)
            ok = values_match(ref_v, api_v, metric)
            note = ""
            if not ok and metric in ("Creator-attributed GMV", "Attributed orders", "Creator-attributed items sold"):
                note = "Shop Analytics 归因口径可能与联盟中心略有差异"
            elif not ok and metric in ("Refunds", "Items refunded"):
                note = "退款字段分析 API 暂无，需 return_refund 或平台导出"
            elif not ok and metric == "Est. commission":
                note = "佣金按 estimated_paid_commission 汇总"
            elif not ok and metric == "Samples shipped":
                note = "联盟订单 API 无寄样字段"
            row = [pid, name, metric, ref_v, api_v, "是" if ok else "否", note]
            ws.append(row)
            total_checks += 1
            if ok:
                match_count += 1
            else:
                for col in range(1, 8):
                    ws.cell(row=ws.max_row, column=col).fill = red

    extra_ids = api_ids - ref_ids
    if extra_ids:
        ws2 = wb.create_sheet("API多出行")
        ws2.append(["Product ID", "Creator-attributed GMV", "说明"])
        for pid in sorted(extra_ids):
            api_dict = dict(zip(EXCEL_HEADERS, api_by_id[pid]))
            ws2.append([pid, api_dict.get("Creator-attributed GMV"), "参考Excel中无此行"])

    ws3 = wb.create_sheet("汇总")
    ws3.append(["项目", "值"])
    ws3.append(["参考文件", str(ref_path)])
    ws3.append(["参考商品数", len(ref_ids)])
    ws3.append(["API商品数(有联盟活动)", len(api_ids)])
    ws3.append(["可比字段数", total_checks])
    ws3.append(["一致字段数", match_count])
    rate = f"{100 * match_count / total_checks:.1f}%" if total_checks else "N/A"
    ws3.append(["一致率", rate])
    ws3.append(["参考有/API无", missing_api])
    ws3.append(["API有/参考无", len(extra_ids)])
    ws3.append(
        [
            "备注",
            "GMV/订单/件数来自 Shop Analytics affiliate_total_performance；"
            "预计佣金来自 affiliate_seller/202410/orders/search；"
            "退款/寄样 API 暂无对应字段",
        ]
    )

    ws4 = wb.create_sheet("API导出副本")
    ws4.append(EXCEL_HEADERS)
    for row in api_rows:
        ws4.append(row)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {
        "ref_count": len(ref_ids),
        "api_count": len(api_ids),
        "match_count": match_count,
        "total_checks": total_checks,
        "missing_api": missing_api,
        "extra_api": len(extra_ids),
    }


def setup_client(shop: str) -> tuple[TikTokShopClient, str, str, Path]:
    config = init_shop_config(ENV_ROOT, ["--shop", shop])
    client = TikTokShopClient(cfg("TTS_APP_KEY"), cfg("TTS_APP_SECRET"))
    client.config_path = config
    token = get_shop_token(client, config)
    if not token:
        raise RuntimeError("无法获取 access_token，请先授权店铺")
    cipher = cfg("TTS_SHOP_CIPHER")
    if not cipher:
        raise RuntimeError("缺少 TTS_SHOP_CIPHER")
    return client, token, cipher, config


def main() -> int:
    ap = argparse.ArgumentParser(description="联盟产品佣金 Excel 导出")
    ap.add_argument("--shop", "-s", default="TK1PH")
    ap.add_argument("--start", default="", help="开始日期（含）；默认前两天")
    ap.add_argument("--end", default="", help="结束日期（含）；默认前两天")
    ap.add_argument(
        "--compare",
        metavar="REF_XLSX",
        help="与联盟中心导出的参考 Excel 对比",
    )
    ap.add_argument(
        "--output-dir",
        default=str(Path.home() / "Desktop"),
        help="输出目录，默认桌面",
    )
    ap.add_argument(
        "--suffix",
        default="",
        help="文件名附加标识，如 联盟订单 → TIKTOK1号店PH_产品佣金_联盟订单_2026-06-16_2026-06-16.xlsx",
    )
    args = ap.parse_args()
    if not args.start or not args.end:
        from shop_tz import default_export_inclusive_range, get_shop_tz, infer_shop_region_from_cfg

        config_pre = init_shop_config(ENV_ROOT, ["--shop", args.shop])
        region = infer_shop_region_from_cfg(config_pre)
        day, _ = default_export_inclusive_range(get_shop_tz(region))
        args.start = args.start or day
        args.end = args.end or day

    client, token, cipher, config = setup_client(args.shop)
    export_tag = cfg("TTS_EXPORT_SHOP_TAG") or args.shop
    single_day = args.start[:10] == args.end[:10]
    region = infer_region()

    print("=" * 60)
    print(f"联盟产品佣金 | {export_tag} | {config.name}")
    print(f"日期: {args.start} ~ {args.end} ({REGION_IANA.get(region, 'Asia/Manila')})")
    print("GMV/订单: Shop Analytics | 佣金: 联盟订单 API")
    print("权限: data.shop_analytics.public.read + seller.affiliate_collaboration.read")
    print("=" * 60)

    affiliate_orders = fetch_all_affiliate_orders(
        client, token, cipher, args.start, args.end, region
    )
    order_by_pid = aggregate_affiliate_orders(affiliate_orders)
    print(f"联盟订单数: {len(affiliate_orders)} | 涉及商品: {len(order_by_pid)}")

    products = fetch_all_product_performance(client, token, cipher, args.start, args.end)
    print(f"分析 API 商品数: {len(products)}")

    meta_cache: dict[str, dict] = {}
    rows: list[list] = []
    seen_pids: set[str] = set()
    for item in products:
        pid = str(item.get("id") or "")
        if not pid:
            continue
        seen_pids.add(pid)
        meta = load_product_meta(client, token, cipher, pid, meta_cache)
        rows.append(
            build_row(
                item,
                meta,
                single_day=single_day,
                order_agg=order_by_pid.get(pid),
            )
        )

    for pid, agg in order_by_pid.items():
        if pid in seen_pids or agg.get("est_commission", 0) <= 0:
            continue
        meta = load_product_meta(client, token, cipher, pid, meta_cache)
        rows.append(
            [
                meta.get("title", ""),
                pid,
                meta.get("category", ""),
                0,
                0,
                "0",
                "0",
                0,
                0,
                0,
                0,
                "",
                "",
                0,
                0,
                round(float(agg.get("est_commission") or 0), 2),
                "",
            ]
        )

    rows = filter_rows_like_affiliate_center(rows)
    print(f"过滤后有联盟活动商品: {len(rows)}")

    out_dir = Path(args.output_dir)
    tag = f"_{args.suffix.strip()}" if args.suffix.strip() else "_API"
    stem = f"{export_tag}_产品佣金{tag}_{args.start}_{args.end}"
    api_path = out_dir / f"{stem}.xlsx"
    export_excel(rows, api_path)
    print(f"已导出: {api_path}")

    api_by_id = {str(r[1]): r for r in rows}

    if args.compare:
        ref_path = Path(args.compare)
        if not ref_path.exists():
            print(f"参考文件不存在: {ref_path}")
            return 1
        cmp_tag = f"_{args.suffix.strip()}" if args.suffix.strip() else ""
        cmp_path = out_dir / f"{export_tag}_产品佣金{cmp_tag}_对比_{args.start}.xlsx"
        stats = write_comparison(ref_path, rows, cmp_path, api_by_id)
        print(f"已生成对比: {cmp_path}")
        if stats["total_checks"]:
            print(
                f"一致率: {stats['match_count']}/{stats['total_checks']} "
                f"({100 * stats['match_count'] / stats['total_checks']:.1f}%)"
            )
        print(f"参考商品 {stats['ref_count']} | API商品 {stats['api_count']}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
