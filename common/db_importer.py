# -*- coding: utf-8 -*-
"""将店铺分析 Excel/缓存导入 MySQL product/sku 报表表。"""
from __future__ import annotations

import json
import re
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from common.cross_border_shop import is_primary_shop_for_import
from common.db_config import load_db_config
from common.file_importer import analytics_cache_filename, local_cache_filename

PRODUCT_HEADERS = {
    "ID": "product_id",
    "商品ID": "product_id",
    "商品": "product_name",
    "状态": "status",
    "GMV": "total_gmv",
    "成交件数": "total_items_sold",
    "订单数": "total_orders",
    "直播归因 GMV": "live_gmv",
    "直播归因成交件数": "live_items_sold",
    "直播曝光次数": "live_impressions",
    "直播的页面浏览次数": "live_page_views",
    "直播的去重页面浏览次数": "live_unique_page_views",
    "直播去重商品客户数": "live_unique_customers",
    "直播点击率": "live_ctr",
    "直播转化率": "live_cvr",
    "视频归因 GMV": "video_gmv",
    "视频归因成交件数": "video_items_sold",
    "视频曝光次数": "video_impressions",
    "来自视频的页面浏览次数": "video_page_views",
    "来自视频的去重页面浏览次数": "video_unique_page_views",
    "视频去重商品客户数": "video_unique_customers",
    "视频点击率": "video_ctr",
    "视频转化率": "video_cvr",
    "商品卡归因 GMV": "card_gmv",
    "商品卡归因成交件数": "card_items_sold",
    "商品卡曝光次数": "card_impressions",
    "商品卡的页面浏览次数": "card_page_views",
    "商品卡的去重页面浏览次数": "card_unique_page_views",
    "商品卡去重客户数": "card_unique_customers",
    "商品卡点击率": "card_ctr",
    "商品卡转化率": "card_cvr",
}

SKU_HEADERS = {
    "SKU ID": "sku_id",
    "SKU编号": "sku_id",
    "Product ID": "product_id",
    "商品ID": "product_id",
    "商品": "sku_name",
    "状态": "status",
    "GMV": "gmv",
    "SKU 订单数": "orders",
    "SKU订单数": "orders",
    "成交件数": "items_sold",
}

PRODUCT_DECIMAL_FIELDS = {
    "total_gmv", "live_gmv", "video_gmv", "card_gmv",
    "live_ctr", "live_cvr", "video_ctr", "video_cvr", "card_ctr", "card_cvr",
}
PRODUCT_INT_FIELDS = {
    "total_items_sold", "total_orders", "live_items_sold", "live_impressions",
    "live_page_views", "live_unique_page_views", "live_unique_customers",
    "video_items_sold", "video_impressions", "video_page_views",
    "video_unique_page_views", "video_unique_customers",
    "card_items_sold", "card_impressions", "card_page_views",
    "card_unique_page_views", "card_unique_customers",
}
SKU_DECIMAL_FIELDS = {"gmv"}
SKU_INT_FIELDS = {"orders", "items_sold"}

_ZERO_SKU_ID = "__ZERO_SKU__"


def _load_cache_payload(
    *,
    z_cache: Path | None = None,
    local_cache: Path | None = None,
    xlsx_path: Path | None = None,
) -> tuple[list[str], list[list[Any]], bool]:
    for cache_file in (z_cache, local_cache):
        if cache_file and cache_file.exists():
            payload = json.loads(cache_file.read_text(encoding="utf-8"))
            return (
                payload.get("headers", []) or [],
                payload.get("rows", []) or [],
                bool(payload.get("zero_data")),
            )
    headers, rows = _load_table_data(
        z_cache=None,
        local_cache=None,
        xlsx_path=xlsx_path,
    )
    return headers, rows, False


def _zero_sku_record(
    *, data_time: str, shop_abbr: str, shop_name: str, site: str
) -> dict[str, Any]:
    return {
        "data_time": data_time,
        "shop_name": shop_name,
        "shop_abbr": shop_abbr,
        "site": site,
        "sku_id": _ZERO_SKU_ID,
        "product_id": _ZERO_SKU_ID,
        "sku_name": "当日零SKU",
        "status": "零数据",
        "gmv": 0.0,
        "orders": 0,
        "items_sold": 0,
    }


def report_date(start: str, api_end: str) -> str:
    """统计日期 = API 区间的展示日（通常等于 start）。"""
    if start:
        return start
    return (date.fromisoformat(api_end) - timedelta(days=1)).isoformat()


def site_from_shop_abbr(shop_abbr: str, site_map: dict[str, str]) -> str:
    abbr = shop_abbr.upper()
    for code in sorted(site_map.keys(), key=len, reverse=True):
        if abbr.endswith(code) or re.search(rf"{code}\d*$", abbr):
            return site_map[code]
    for code, name in (("PH", "菲律宾"), ("TH", "泰国"), ("MY", "马来")):
        if code in abbr:
            return site_map.get(code, name)
    return site_map.get("PH", "菲律宾")


def _parse_decimal(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "").replace("₱", "").replace("PHP", "")
    if s.endswith("%"):
        s = s[:-1].strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip().replace(",", "")
    try:
        return int(float(s))
    except ValueError:
        return 0


def _load_table_data(
    *,
    z_cache: Path | None = None,
    local_cache: Path | None = None,
    xlsx_path: Path | None = None,
) -> tuple[list[str], list[list[Any]]]:
    for cache_file in (z_cache, local_cache):
        if cache_file and cache_file.exists():
            payload = json.loads(cache_file.read_text(encoding="utf-8"))
            return payload.get("headers", []), payload.get("rows", [])

    if xlsx_path and xlsx_path.exists():
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return [], []
        headers = [str(c) if c is not None else "" for c in rows[0]]
        body = [list(r) for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        return headers, body

    return [], []


def _row_dict(headers: list[str], row: list[Any], header_map: dict[str, str]) -> dict[str, Any]:
    idx = {h: i for i, h in enumerate(headers)}
    out: dict[str, Any] = {}
    for src, dst in header_map.items():
        if src not in idx:
            continue
        val = row[idx[src]]
        out[dst] = "" if val is None else val
    return out


def _build_product_rows(
    headers: list[str],
    rows: list[list[Any]],
    *,
    data_time: str,
    shop_abbr: str,
    shop_name: str,
    site: str,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in rows:
        item = _row_dict(headers, row, PRODUCT_HEADERS)
        product_id = str(item.get("product_id", "")).strip()
        if not product_id:
            continue
        rec = {
            "data_time": data_time,
            "shop_name": shop_name,
            "shop_abbr": shop_abbr,
            "site": site,
            "product_id": product_id,
            "product_name": str(item.get("product_name", "") or ""),
            "status": str(item.get("status", "") or ""),
        }
        for field in PRODUCT_DECIMAL_FIELDS:
            rec[field] = _parse_decimal(item.get(field))
        for field in PRODUCT_INT_FIELDS:
            rec[field] = _parse_int(item.get(field))
        records.append(rec)
    return records


def _build_sku_rows(
    headers: list[str],
    rows: list[list[Any]],
    *,
    data_time: str,
    shop_abbr: str,
    shop_name: str,
    site: str,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in rows:
        item = _row_dict(headers, row, SKU_HEADERS)
        sku_id = str(item.get("sku_id", "")).strip()
        product_id = str(item.get("product_id", "")).strip()
        if not sku_id or not product_id:
            continue
        rec = {
            "data_time": data_time,
            "shop_name": shop_name,
            "shop_abbr": shop_abbr,
            "site": site,
            "sku_id": sku_id,
            "product_id": product_id,
            "sku_name": str(item.get("sku_name", "") or ""),
            "status": str(item.get("status", "") or ""),
        }
        for field in SKU_DECIMAL_FIELDS:
            rec[field] = _parse_decimal(item.get(field))
        for field in SKU_INT_FIELDS:
            rec[field] = _parse_int(item.get(field))
        records.append(rec)
    return records


def _upsert_many(conn, table: str, rows: list[dict[str, Any]], unique_cols: list[str]) -> tuple[int, int]:
    if not rows:
        return 0, 0

    cols = list(rows[0].keys())
    placeholders = ", ".join(["%s"] * len(cols))
    col_sql = ", ".join(f"`{c}`" for c in cols)
    update_cols = [c for c in cols if c not in unique_cols]
    update_sql = ", ".join(f"`{c}`=VALUES(`{c}`)" for c in update_cols)
    sql = (
        f"INSERT INTO `{table}` ({col_sql}) VALUES ({placeholders}) "
        f"ON DUPLICATE KEY UPDATE {update_sql}"
    )

    values = [tuple(row[c] for c in cols) for row in rows]
    with conn.cursor() as cur:
        cur.executemany(sql, values)
    conn.commit()
    return len(rows), 0


def import_shop_analytics_to_db(
    *,
    shop_key: str,
    shop_name: str,
    export_tag: str,
    logs_dir: Path,
    start: str,
    api_end: str,
    data_types: set[str],
    product_xlsx: Path | None = None,
    sku_xlsx: Path | None = None,
    cache_dirs: dict[str, Path] | None = None,
    db_ini: Path | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    cfg = load_db_config(db_ini)
    data_time = report_date(start, api_end)
    site = site_from_shop_abbr(shop_key, cfg["site_map"])
    shop_display = export_tag or shop_name or shop_key

    ok_primary, primary_msg = is_primary_shop_for_import(shop_key)
    if not ok_primary:
        return {
            "shop": shop_key,
            "data_time": data_time,
            "site": site,
            "product_rows": 0,
            "sku_rows": 0,
            "ok": True,
            "messages": [f"[跳过入库] {primary_msg}"],
            "skipped_non_primary": True,
        }

    result = {
        "shop": shop_key,
        "data_time": data_time,
        "site": site,
        "product_rows": 0,
        "sku_rows": 0,
        "ok": True,
        "messages": [],
    }

    product_records: list[dict[str, Any]] = []
    sku_records: list[dict[str, Any]] = []
    cache_dirs = cache_dirs or {}
    logs_shop = logs_dir / shop_key

    if "product" in data_types:
        z_product_cache = None
        if cache_dirs.get("产品_JSON目录"):
            z_product_cache = (
                cache_dirs["产品_JSON目录"]
                / analytics_cache_filename(export_tag, "product", start, api_end)
            )
        headers, rows = _load_table_data(
            z_cache=z_product_cache,
            local_cache=logs_shop / local_cache_filename("product", start, api_end),
            xlsx_path=product_xlsx,
        )
        product_records = _build_product_rows(
            headers, rows,
            data_time=data_time,
            shop_abbr=shop_key,
            shop_name=shop_display,
            site=site,
        )
        result["product_rows"] = len(product_records)

    if "sku" in data_types:
        z_sku_cache = None
        if cache_dirs.get("SKU_JSON目录"):
            z_sku_cache = (
                cache_dirs["SKU_JSON目录"]
                / analytics_cache_filename(export_tag, "sku", start, api_end)
            )
        headers, rows, sku_zero_data = _load_cache_payload(
            z_cache=z_sku_cache,
            local_cache=logs_shop / local_cache_filename("sku", start, api_end),
            xlsx_path=sku_xlsx,
        )
        if rows:
            sku_records = _build_sku_rows(
                headers,
                rows,
                data_time=data_time,
                shop_abbr=shop_key,
                shop_name=shop_display,
                site=site,
            )
        elif sku_zero_data:
            sku_records = [
                _zero_sku_record(
                    data_time=data_time,
                    shop_abbr=shop_key,
                    shop_name=shop_display,
                    site=site,
                )
            ]
        result["sku_rows"] = len(sku_records)

    if dry_run:
        result["messages"].append(
            f"[dry-run] {shop_key} {data_time} product={len(product_records)} sku={len(sku_records)}"
        )
        return result

    if not product_records and not sku_records:
        result["messages"].append(f"[跳过] {shop_key} 无 product/sku 数据")
        return result

    if sku_records and sku_records[0].get("sku_id") == _ZERO_SKU_ID:
        result["messages"].append(
            f"[标记] {shop_key} {data_time} 当日零SKU（已写入占位行，查缺不再报缺失）"
        )

    try:
        import pymysql
    except ImportError as exc:
        raise RuntimeError("请先安装 pymysql: pip install pymysql") from exc

    conn = pymysql.connect(
        host=cfg["host"],
        port=cfg["port"],
        user=cfg["user"],
        password=cfg["password"],
        database=cfg["database"],
        charset=cfg["charset"],
    )
    try:
        if product_records:
            n, _ = _upsert_many(
                conn,
                cfg["product_table"],
                product_records,
                ["data_time", "shop_abbr", "site", "product_id"],
            )
            result["messages"].append(f"[DB] product 写入 {n} 条 -> {cfg['product_table']}")

        if sku_records:
            n, _ = _upsert_many(
                conn,
                cfg["sku_table"],
                sku_records,
                ["data_time", "shop_abbr", "site", "sku_id"],
            )
            result["messages"].append(f"[DB] sku 写入 {n} 条 -> {cfg['sku_table']}")
    finally:
        conn.close()

    return result
